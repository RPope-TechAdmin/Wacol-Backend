import os
import time
import logging
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Iterable, Tuple
import azure.functions as func
import pyodbc
import openpyxl

# ========= CONFIG =========
# Read from Function App settings (Configuration → Application settings)
SQL_SERVER = "purenvqld.database.windows.net"
SQL_DATABASE = "LaboratoryQLD"
SQL_USERNAME = "reportabledatadownloader"
SQL_PASSWORD = "Rep0r7D47aD0wn"

# Map HTML group ids → fully-qualified table names (include schema!)
GROUP_TO_TABLE: Dict[str, str] = {
    "fixtotal": "[Wacol].[Fixation]",     # Fixation Totals
    "twint": "[Wacol].[Trade Waste]",     # Internal Trade Waste Testing
    "sw": "[Wacol].[Stormwater]",     # Stormwater Testing
    }

# Whitelist analyte columns per table to avoid SQL injection via column names.
# ⚠️ 
ALLOWED_COLUMNS: Dict[str, Iterable[str]] = {    
    "[Wacol].[Fixation]": {
        "File","Sample Date","Sample Name","Moisture Content","4.4`-DDD","4.4`-DDE","4.4`-DDT","Aldrin","alpha-BHC","alpha-Endosulfan","Azinphos Methyl","beta-BHC","beta-Endosulfan","Bromophos-ethyl","Carbophenothion","Chlorfenvinphos","Chlorpyrifos","Chlorpyrifos-methyl","cis-Chlordane"
        ,"delta-BHC","Demeton-S-methyl","Diazinon","Dichlorvos","Dieldrin","Dimethoate","Endosulfan (sum)","Endosulfan sulfate","Endrin","Endrin aldehyde","Endrin ketone","Ethion","Fenamiphos","Fenthion","gamma-BHC - (Lindane)","Heptachlor","Heptachlor epoxide","Hexachlorobenzene (HCB)","Malathion"
        ,"Methoxychlor","Monocrotophos","Parathion","Parathion-methyl","Pirimphos-ethyl","Prothiofos","Sum of Aldrin + Dieldrin","Sum of DDD + DDE + DDT","Total Chlordane (sum)","trans-Chlordane",">C10 - C16 Fraction",">C10 - C16 Fraction minus Naphthalene (F2)",">C10 - C40 Fraction (sum)"
        ,">C16 - C34 Fraction",">C34 - C40 Fraction","C10 - C14 Fraction","C10 - C36 Fraction (sum)","C15 - C28 Fraction","C29 - C36 Fraction","Benzene","C6 - C10 Fraction","C6 - C10 Fraction minus BTEX (F1)","C6 - C9 Fraction","Ethylbenzene","meta- & para-Xylene","Naphthalene"
        ,"ortho-Xylene","Sum of BTEX","Toluene","Total Xylenes","2.4.5-Trichlorophenol","2.4.6-Trichlorophenol","2.4-Dichlorophenol","2.4-Dimethylphenol","2.6-Dichlorophenol","2-Chlorophenol","2-Methylphenol","2-Nitrophenol","3- & 4-Methylphenol","4-Chloro-3-methylphenol","Acenaphthene"
        ,"Acenaphthylene","Anthracene","Benz(a)anthracene","Benzo(a)pyrene","Benzo(a)pyrene TEQ (half LOR)","Benzo(a)pyrene TEQ (LOR)","Benzo(a)pyrene TEQ (zero)","Benzo(b+j)fluoranthene","Benzo(g.h.i)perylene","Benzo(k)fluoranthene","Chrysene","Dibenz(a.h)anthracene","Fluoranthene","Fluorene"
        ,"Indeno(1.2.3.cd)pyrene","PAH Naphthalene","Pentachlorophenol","Phenanthrene","Phenol","Pyrene","Sum of polycyclic aromatic hydrocarbons","Antimony","Arsenic","Barium","Beryllium","Boron","Cadmium","Chromium","Cobalt","Copper","Lead","Manganese","Molybdenum","Nickel","Selenium"
        ,"Tin","Zinc","Mercury","After HCl pH","Extraction Fluid Number","Final pH","Initial pH","ZHE Extraction Fluid Number","TCLP 4.4`-DDD","TCLP 4.4`-DDE","TCLP 4.4`-DDT","TCLP Aldrin","TCLP alpha-BHC","TCLP alpha-Endosulfan","TCLP Azinphos Methyl","TCLP beta-BHC","TCLP beta-Endosulfan"
        ,"TCLP Bromophos-ethyl","TCLP Carbophenothion","TCLP Chlorfenvinphos","TCLP Chlorpyrifos","TCLP Chlorpyrifos-methyl","TCLP cis-Chlordane","TCLP delta-BHC","TCLP Demeton-S-methyl","TCLP Diazinon","TCLP Dichlorvos","TCLP Dieldrin","TCLP Dimethoate","TCLP Endosulfan sulfate","TCLP Endrin"
        ,"TCLP Endrin aldehyde","TCLP Endrin ketone","TCLP Ethion","TCLP Fenamiphos","TCLP Fenthion","TCLP gamma-BHC - (Lindane)","TCLP Heptachlor","TCLP Heptachlor epoxide","TCLP Hexachlorobenzene (HCB)","TCLP Malathion","TCLP Methoxychlor","TCLP Monocrotophos","TCLP Parathion","TCLP Parathion-methyl"
        ,"TCLP Pirimphos-ethyl","TCLP Prothiofos","TCLP Sum of Aldrin + Dieldrin","TCLP Sum of DDD + DDE + DDT","TCLP Total Chlordane (sum)","TCLP trans-Chlordane","TCLP >C10 - C16 Fraction","TCLP >C10 - C16 Fraction minus Naphthalene (F2)","TCLP >C10 - C40 Fraction (sum)","TCLP >C16 - C34 Fraction"
        ,"TCLP >C34 - C40 Fraction","TCLP C10 - C14 Fraction","TCLP C10 - C36 Fraction (sum)","TCLP C15 - C28 Fraction","TCLP C29 - C36 Fraction","TCLP Benzene","TCLP C6 - C10 Fraction","TCLP C6 - C10 Fraction minus BTEX (F1)","TCLP C6 - C9 Fraction","TCLP Ethylbenzene","TCLP meta- & para-Xylene"
        ,"TCLP Naphthalene","TCLP ortho-Xylene","TCLP Sum of BTEX","TCLP Toluene","TCLP Total Xylenes","TCLP 2.4.5-Trichlorophenol","TCLP 2.4.6-Trichlorophenol","TCLP 2.4-Dichlorophenol","TCLP 2.4-Dimethylphenol","TCLP 2.6-Dichlorophenol","TCLP 2-Chlorophenol","TCLP 2-Methylphenol","TCLP 2-Nitrophenol"
        ,"TCLP 3- & 4-Methylphenol","TCLP 4-Chloro-3-methylphenol","TCLP Acenaphthene","TCLP Acenaphthylene","TCLP Anthracene","TCLP Benz(a)anthracene","TCLP Benzo(a)pyrene","TCLP Benzo(a)pyrene TEQ (zero)","TCLP Benzo(b+j)fluoranthene","TCLP Benzo(g.h.i)perylene","TCLP Benzo(k)fluoranthene","TCLP Chrysene"
        ,"TCLP Dibenz(a.h)anthracene","TCLP Fluoranthene","TCLP Fluorene","TCLP Indeno(1.2.3.cd)pyrene","TCLP PAH Naphthalene","TCLP Pentachlorophenol","TCLP Phenanthrene","TCLP Phenol","TCLP Pyrene","TCLP Sum of polycyclic aromatic hydrocarbons","TCLP Antimony","TCLP Arsenic","TCLP Barium","TCLP Beryllium"
        ,"TCLP Boron","TCLP Cadmium","TCLP Chromium","TCLP Cobalt","TCLP Copper","TCLP Lead","TCLP Manganese","TCLP Molybdenum","TCLP Nickel","TCLP Selenium","TCLP Tin","TCLP Zinc","TCLP Mercury"
    },
    "[Wacol].[Trade Waste]": {
        "File","Sample Date","Sample Name","4.4`-DDD","4.4`-DDE","4.4`-DDT","Aldrin","alpha-BHC","alpha-Endosulfan","Azinphos Methyl","beta-BHC","beta-Endosulfan","Bromophos-ethyl","Carbophenothion","Chlorfenvinphos","Chlorpyrifos","Chlorpyrifos-methyl","cis-Chlordane","delta-BHC","Demeton-S-methyl","Diazinon"
        ,"Dichlorvos","Dieldrin","Dimethoate","Endosulfan sulfate","Endrin","Endrin aldehyde","Endrin ketone","Ethion","Fenamiphos","Fenthion","gamma-BHC - (Lindane)","Heptachlor","Heptachlor epoxide","Hexachlorobenzene (HCB)","Malathion","Methoxychlor","Monocrotophos","Parathion","Parathion-methyl"
        ,"Pirimphos-ethyl","Prothiofos","Sum of Aldrin + Dieldrin","Sum of DDD + DDE + DDT","Total Chlordane (sum)","trans-Chlordane","Mercury","Arsenic","Cadmium","Chromium","Copper","Lead","Nickel","Zinc",">C10 - C16 Fraction",">C10 - C16 Fraction minus Naphthalene (F2)",">C10 - C40 Fraction (sum)"
        ,">C16 - C34 Fraction",">C34 - C40 Fraction","C10 - C14 Fraction","C10 - C36 Fraction (sum)","C15 - C28 Fraction","C29 - C36 Fraction","Benzene","C6 - C10 Fraction","C6 - C10 Fraction minus BTEX (F1)","C6 - C9 Fraction","Ethylbenzene","meta- & para-Xylene","Naphthalene","ortho-Xylene","Sum of BTEX"
        ,"Toluene","Total Xylenes","AMPA","Glyphosate","2.4.5-T","2.4.6-T","2.4-D","2.4-DB","2.4-DP","2.6-D","4-Chlorophenoxy acetic acid","Clopyralid","Dicamba","Fluroxypyr","MCPA","MCPB","Mecoprop","Picloram","Silvex (2.4.5-TP/Fenoprop)","Triclopyr","pH Value","Total Nitrogen as N","Total Kjeldahl Nitrogen as N"
        ,"Nitrite + Nitrate as N","Nitrate as N","Nitrite as N","Ammonia as N","Total Phosphorus as P","Biochemical Oxygen Demand","Chemical Oxygen Demand","Suspended Solids (SS)","Total Organic Carbon","Dilution Factor"
    },
    "[Wacol].[Stormwater]": {
        "File","Sample Date","Sample Name",">C10 - C16 Fraction",">C10 - C16 Fraction minus Naphthalene (F2)",">C10 - C40 Fraction (sum)",">C16 - C34 Fraction",">C34 - C40 Fraction","C10 - C14 Fraction","C10 - C36 Fraction (sum)","C15 - C28 Fraction","C29 - C36 Fraction","Benzene","C6 - C10 Fraction","C6 - C10 Fraction minus BTEX (F1)"
        ,"C6 - C9 Fraction","Ethylbenzene","meta- & para-Xylene","Naphthalene","ortho-Xylene","Sum of BTEX","Toluene","Total Xylenes","pH Value","Electrical Conductivity @ 25°C","Suspended Solids (SS)","Total Organic Carbon","Turbidity"
    }
}

ID_COLUMNS = ["File","Sample Date", "Sample Name"]

# Non-analyte identifier columns you always want back
TABLE_ID_COLUMNS: Dict[str, List[str]] = {
    "[Wacol].[Fixation]": ["File","Sample Date"],
    "[Wacol].[Trade Waste]": ["File Name","Sample Date"],
    "[Wacol].[Stormwater]": ["File Name","Sample Date"]
}

# ========= DB CONNECT =========
def connect_with_fallback(timeout_seconds: int = 60) -> pyodbc.Connection:
    """
    Try ODBC Driver 18 then 17. Increase Connection Timeout and retry a few times
    (useful if Azure SQL Serverless is resuming).
    """
    drivers = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"]
    last_exc = None

    for driver in drivers:
        conn_str = (
            f"Driver={{{driver}}};"
            f"Server=tcp:{SQL_SERVER},1433;"
            f"Database={SQL_DATABASE};"
            f"Uid={SQL_USERNAME};"
            f"Pwd={SQL_PASSWORD};"
            "Encrypt=yes;"
            "TrustServerCertificate=no;"
            f"Connection Timeout={timeout_seconds};"
        )
        for attempt in range(3):
            try:
                logging.info("Attemting DB Connection")
                return pyodbc.connect(conn_str)
            except Exception as e:
                last_exc = e
                logging.warning(f"Connect attempt {attempt+1}/3 with {driver} failed: {e}")
                time.sleep(3)
    # If we get here, all attempts failed
    raise last_exc

# ========= HELPERS =========
def normalize_payload(data) -> Dict[str, List[str]]:
    """
    Support two payload shapes:
      A) { "selections": { "external": ["arsenic", ...], "pfastopa": ["pfos"] }, "startDate": "...", "endDate": "..." }
      B) { "selections": [ { "table": "Jackson.DSExt", "analyte":"arsenic" }, ... ], ... }
    Returns a dict { group_key_or_table: [analytes...] } keyed by group-id (preferred) if possible.
    """
    sel = data.get("selections", [])
    if isinstance(sel, dict):
        # Already grouped by HTML group id
        return {k: list(v) for k, v in sel.items() if v}
    elif isinstance(sel, list):
        grouped: Dict[str, List[str]] = {}
        for item in sel:
            table = item.get("table")
            analyte = item.get("analyte")
            if not table or not analyte:
                continue
            # Try to reverse-map table back to group id; if not found, use the table name as the key
            group_key = None
            for g, t in GROUP_TO_TABLE.items():
                if t.lower() == table.lower():
                    group_key = g
                    break
            key = group_key if group_key else table
            grouped.setdefault(key, []).append(analyte)
        return grouped
    else:
        return {}

def whitelist_columns(table: str, requested: Iterable[str]) -> List[str]:
    allowed = set(ALLOWED_COLUMNS.get(table, []))
    return [c for c in requested if c in allowed]

def build_select_sql(table: str, analyte_cols: List[str]) -> str:
    # Fall back to global ID_COLUMNS if table not in map
    id_cols = TABLE_ID_COLUMNS.get(table, ID_COLUMNS)

    selected_cols = id_cols + analyte_cols
    cols_sql = ", ".join(f"[{c}]" for c in selected_cols)  # bracket-quote identifiers
    return f"SELECT {cols_sql} FROM {table} WHERE [Sample Date] BETWEEN ? AND ?"
    
def safe_sheet_name(name: str) -> str:
    # Excel sheet name: max 31 chars, no []:*?/\
    bad = '[]:*?/\\'
    for ch in bad:
        name = name.replace(ch, "-")
    return (name or "Sheet")[:31]

# ========= MAIN FUNCTION =========

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Azure function 'lab-data-download' has been triggered.")
    logging.info("Processing /download-excel request")
    try:
        data = req.get_json()
    except ValueError:
        logging.error("Request body is not valid JSON.")
        return func.HttpResponse("Invalid JSON body.", status_code=400)

    start_date = data.get("startDate")
    end_date = data.get("endDate")
    if not start_date or not end_date:
        return func.HttpResponse("Both startDate and endDate are required.", status_code=400)

    logging.debug(f"Received request with start date: {start_date}, end date: {end_date}")

    grouped = normalize_payload(data)
    if not grouped:
        return func.HttpResponse("No analytes selected.", status_code=400)
    
    logging.debug(f"Normalized payload selections: {grouped}")

    # Open workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    try:
        conn = connect_with_fallback(timeout_seconds=60)
        cursor = conn.cursor()
        logging.info("Successfully connected to the database.")

        any_rows_written = False
        sheets_created = 0

        for group_key, analytes in grouped.items():
            table = GROUP_TO_TABLE.get(group_key, group_key)
            if table not in ALLOWED_COLUMNS:
                logging.warning(f"Skipping unknown/unauthorized table: {table}")
                continue

            logging.debug(f"Processing group '{group_key}' for table '{table}'.")

            analyte_cols = whitelist_columns(table, analytes)
            if not analyte_cols:
                logging.info(f"No valid analyte columns for {table}, requested: {analytes}")
                continue
            
            logging.debug(f"Whitelisted columns for {table}: {analyte_cols}")

            sql = build_select_sql(table, analyte_cols)
            logging.info(f"Running query for {table}: {sql}")
            cursor.execute(sql, (start_date, end_date))
            rows = cursor.fetchall()
            columns = [d[0] for d in cursor.description]
            logging.info(f"Query for {table} returned {len(rows)} rows.")

            ws = wb.create_sheet(title=safe_sheet_name(group_key))
            sheets_created += 1
            ws.append(columns)

            if rows:
                for row in rows:
                    ws.append(list(row))
                any_rows_written = True
                logging.debug(f"Wrote {len(rows)} data rows to sheet '{ws.title}'.")
            else:
                ws.append(["No data found for this selection."])

        if not wb.worksheets:  # safety: if nothing created
            ws = wb.create_sheet(title="Results")
            ws.append(["No data found at all."])

        logging.info(f"Created {sheets_created} sheets for the Excel file.")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return func.HttpResponse(
            body=output.getvalue(),
            status_code=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logging.error(f"Download error: {e}", exc_info=True)
        return func.HttpResponse(f"Error: {e}", status_code=500)
    finally:
        if 'cursor' in locals() and cursor:
            try: cursor.close()
            except Exception: pass
        if 'conn' in locals() and conn:
            try: conn.close()
            except Exception: pass
        logging.debug("Database connection and cursor closed.")